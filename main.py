import requests
import sys
from datetime import datetime
from openpyxl import load_workbook


FILENAME = "project.xlsx"
url = "https://tradedata.go.kr/cts/hmpg/retrieveTrade.do"
STANDARD_SHEET_NAME = "코드수정(최종)"
BYNATION_SHEET_NAME = "국가별 수출(월별)"
TOTAL = "합계"
TOTAL_ALL = "월별 수출 총액"
TOTAL_ALL_BY_COUNTRY = "수출총액"
FORE = "FORE"
now = datetime.now()
YEAR = now.year
MONTH = now.month

'''
TW=대만,  HK=홍콩
파일에 없으니 추가할것.
'''


def numToChr(num):
    '''
    :param num: 숫자. 열 번호
    :return: 1->A, 2->B, ..., 27->AA, ...의 엑셀의 열 번호를 구현했다.
    '''
    if num == 0:
        return ""
    else:
        return numToChr((num - 1) // 26) + chr(ord("A") + ((num - 1) % 26))


class dataChart:
    standardChart = None            # 양식의 표준이 되는 종합수입차트
    standardExChart=None            # 종합수출차트
    chartByNations = {}
    hsSgns = []                     # 표준양식의 코드모음
    code2row = {}                   # 코드->몇번째 줄인지
    itemInfos = []                  # 리스트. 해당 row가 어떤 코드에 관한 것인지 카테고리 소합계(TOTAL)인지, 아니면 대합계(TOTAL_ALL)인지 나온다.
    nameInfos = []
    country2code = {}               # 미국 -> US의 딕셔너리
    code2country = {}               # US->미국의 딕셔너리
    toDo = "AUTO"                   # 세팅옵션
    addNation = "NO"                # 세팅옵션
    saveNew = "NO"                  # 세팅옵션
    fileTempt = "TEMPT"             # 세팅옵션
    foreYear = 0                  # 표의 예측 날짜
    foreMonth=0
    startYear = 0                   # 표의 첫 날짜(2015)
    startMonth = 0                  # 표의 첫 날짜(1)      -> year*100+month로 requests param에 넣어주면 된다.
    code2chart = {}                 # 국가코드-> 차트 오브젝트
    totalIndexes=[]                 # 부분합의 로우값에 대한 정보.
    worksheetByNation=None

    def __init__(self, worksheet, rownum):
        self.worksheet = worksheet
        self.rownum = rownum
        self.to_standard_row()

    def to_standard_row(self):
        if self.itemInfos:
            nowrow = self.rownum + 3
            flag = False
            for row, (item, name) in enumerate(zip(self.itemInfos[:-1], self.nameInfos),nowrow):
                while item!=(cellval:=(self.worksheet.cell(row, 3).value or self.worksheet.cell(row, 1).value)):
                    flag=True
                    if cellval in self.itemInfos:
                        self.worksheet.insert_rows(row)
                        self.worksheet.cell(row=row, column=2, value=name)
                        if name:
                            self.worksheet.cell(row=row, column=3, value=item)
                        else:
                            self.worksheet.cell(row=row, column=1, value=item)
                        break
                    else:
                        self.worksheet.delete_rows(row)
            if flag:
                self.correct_cell_sum_function_values()

    def correct_cell_sum_function_values(self):
        startColumn=4
        endColumn= (YEAR+1-self.startYear)*12 - self.startMonth + 4 + 1

        beforeSumRow=self.rownum+2
        rows = []
        for nowrow in self.totalIndexes:
            row = self.rownum + 3 + nowrow
            for column in range(startColumn, endColumn):
                corchr = numToChr(column)
                self.worksheet.cell(row=row, column=column, value=f"=SUM({corchr}{beforeSumRow+1}:{corchr}{row-1})")
            beforeSumRow = row
            rows.append(str(beforeSumRow))
        row = self.rownum + 2 + len(self.itemInfos)
        for column in range(startColumn, endColumn):
            corchr=numToChr(column)
            self.worksheet.cell(row=row, column=column, value=f"=SUM({corchr}{f',{corchr}'.join(rows)})")

    @classmethod
    def __get_setting(cls, settings_worksheet):
        '''
        엑셀파일의 세팅시트를 읽어 옵션을 세팅합니다.
        '''
        if settings_worksheet != None:
            cls.toDo = settings_worksheet.cell(1, 2).value or "AUTO"
            cls.addNation = settings_worksheet.cell(2, 2).value or "NO"
            cls.saveNew = settings_worksheet.cell(3, 2).value or "NO"
            cls.fileTempt = settings_worksheet.cell(4, 2).value or "TEMPT"
            CntryCode = settings_worksheet.cell(5, 2).value
            f = open(CntryCode, 'r')
            ll = f.readlines()
            f.close()
            for l in ll:
                l = l.strip().split(",")
                try:
                    cls.country2code[l[-1]] = l[0]
                    cls.code2country[l[0]] = l[-1]
                except:
                    pass
        else:
            cls.toDo = "AUTO"
            cls.addNation = "NO"
            cls.saveNew = "NO"
            cls.fileTempt = f"TEMPT"
            print("WARNING: settings를 찾을 수 없습니다.")
        print(
            f"할 일:{cls.toDo}, 국가 추가:{cls.addNation}, 파일생성:{cls.saveNew}, 임시파일명:{cls.fileTempt}, 국가코드경로:{CntryCode} 옵션으로 진행합니다.")

    @classmethod
    def settings(cls, workbook):
        '''
        workboook을 받아서 기본값을 설정해줍니다.
        '''
        cls.wb = workbook
        try:
            setting_ws = cls.wb["settings"]
        except:
            setting_ws = None
        cls.__get_setting(setting_ws)

    @classmethod
    def create_country_chart(cls, worksheet):
        cls.worksheetByNation=worksheet
        row=1
        gap = len(cls.itemInfos) + 4
        while (v:= worksheet.cell(row, 1).value):
            cls.code2chart[cls.country2code.get(v,v)] = cls(worksheet, row)
            row+=gap

    @classmethod
    def create_standard_chart(cls, worksheet, rownum):
        '''
        기본차트를 생성합니다. 주 양식에 대한 정보가 들어갑니다.
        통합수입표를 주 양식으로 잡으며, 그 아래 표를 통합수출표로 기록합니다.
        :param worksheet: 양식이 위치한 워크시트입니다.
        :param rownum: 양식의 가장 상단의 행번호입니다.
        '''
        cls.standardChart = cls(worksheet, rownum)
        try:
            row = 3
            while True:
                ai = worksheet.cell(rownum + row, 1).value
                bi = worksheet.cell(rownum+row, 2).value
                ci = worksheet.cell(rownum + row, 3).value
                cls.itemInfos.append(ci or ai)
                cls.nameInfos.append(bi)
                if ci:
                    hsSgn = "".join(ci.split("-"))
                    cls.hsSgns.append(hsSgn)
                    cls.code2row[hsSgn] = row
                elif ai==TOTAL:
                    cls.totalIndexes.append(len(cls.itemInfos)-1)
                elif ai == TOTAL_ALL:
                    break
                else:
                    print(f"서식에 문제가 생겼습니다. {row}번 행을 수정 바랍니다.")
                    sys.exit(0)
                row += 1
        except:
            print("ERROR: 양식에 문제가 있습니다. 코드수정(최종) 시트의 양식을 지켜 주시기 바랍니다.")
            input("종료하기 위해서 엔터를 눌러주세요")
            sys.exit(0)
        Y, M = f"{worksheet.cell(rownum + 1, 4).value}".split()
        cls.startYear = int(Y[:-1])
        cls.startMonth = int(M[:-1])
        dy = YEAR - cls.startYear
        dm = MONTH - cls.startMonth
        colnum = dy * 12 + dm + 4
        while colnum > 3:
            if worksheet.cell(rownum, colnum).value == FORE:
                cls.foreMonth = (colnum-4)%12 + 1
                cls.foreYear = (colnum - 4)//12 + 2015
                break
            colnum -= 1
        cls.standardExChart = cls(worksheet,rownum+len(cls.itemInfos)+2+3)
        return cls.standardChart

    @classmethod
    def run(cls):
        '''
        데이터를 다 채우는 실행을 합니다.
        모드에 따라서 다른 행동을 해야합니다.
        :return:
        '''
        cls.__fill_main_page()
        cls.__fill_country_page()

    @classmethod
    def __fill_main_page(cls):
        '''
        코드수정(최종)시트를 채우는 함수.
        cls.standardChart : 수입종합차트
        cls.standardExChart : 수출종합차트
        데이터를 request로 불러와서 _fiillchart를 호출하는 방법으로 채워줍니다.
            ex) cls.standardChart._fillChart(data, "impUsdAmt")
        '''

        frompriod = cls.foreYear*100 + cls.foreMonth
        if cls.toDo=="VALIDATE":
            frompriod =cls.startYear*100+cls.startMonth

        endpriod = YEAR*100 + MONTH
        data = {
            "tradeKind": "ETS_MNK_1020000A",
            "priodKind": "MON",
            "priodFr": frompriod,
            "priodTo": endpriod,
            "statsBase": "acptDd",
            "ttwgTpcd": "1000",
            "showPagingLine": 1000000,
            "sortColumn": "",
            "sortOrder": "",
            "hsSgnGrpCol": "HS10_SGN",
            "hsSgnWhrCol": "HS10_SGN",
            "hsSgn": cls.hsSgns
        }
        datas = requests.post(url=url, data=data).json()

        maxcolnum=0
        for data in datas["items"]:
            maxcolnum=max(maxcolnum,cls.standardChart.__fill_chart(data, "expUsdAmt"))             # 수출
            cls.standardExChart.__fill_chart(data, "impUsdAmt")           # 수입

        if(len(datas["items"])==0):
            print("업데이트할 데이터가 없습니다.")

        if maxcolnum:
            yyyy, mm = frompriod//100, frompriod%100
            dy = yyyy - cls.startYear
            dm = mm - cls.startMonth
            startcolumn = dy * 12 + dm + 4

            for col in range(startcolumn, maxcolnum+1):
                '''
                column의 데이터값에 대해 "=F"로 시작한다면 0으로 바꿀것.
                '''
                cls.standardChart.fillZero(col)
                cls.standardExChart.fillZero(col)
                pass

            beforeForeCol = (cls.foreYear-cls.startYear) * 12 + (cls.foreMonth - cls.startMonth) + 4
            cls.standardChart.worksheet.cell(1,beforeForeCol,value="")
            cls.standardChart.worksheet.cell(1, maxcolnum+1, value=FORE)

            cls.standardExChart.__make_forecast(maxcolnum+1, True)

    def fillZero(self, column):
        for row in range(3, len(self.itemInfos) + 3):
            val=self.worksheet.cell(self.rownum+row, column).value
            if str(val)[:2]== "=F" or val==None:
                self.worksheet.cell(self.rownum+row, column, value=0)

    @classmethod
    def __fill_country_page(cls):
        frompriod = cls.foreYear * 100 + cls.foreMonth
        endpriod = YEAR * 100 + MONTH
        cntyNm = []
        data = {
            "tradeKind": "ETS_MNK_1020000E",
            "priodKind": "MON",
            "priodFr": frompriod,
            "priodTo": endpriod,
            "statsBase": "acptDd",
            "ttwgTpcd": "1000",
            "showPagingLine": 1000000,
            "sortColumn": "",
            "sortOrder": "",
            "hsSgnGrpCol": "HS10_SGN",
            "hsSgnWhrCol": "HS10_SGN",
            "hsSgn": cls.hsSgns,
            "cntyNm": cntyNm
        }
        maxcolnum=0
        datas = requests.post(url=url, data=data).json()
        to_add = set()
        for data in datas["items"]:
            if data["hsSgn"].strip():
                if data["cntyCd"] in cls.code2chart:
                    maxcolnum = max(cls.code2chart[data["cntyCd"]].__fill_chart(data, "expUsdAmt"), maxcolnum)
                else:
                    if cls.addNation=="NO":
                        to_add.add(data['cntyCd'])
                    else:
                        cls.create_chart(data['cntyCd'], cls.worksheetByNation)
                        maxcolnum = max(cls.code2chart[data["cntyCd"]].__fill_chart(data, "expUsdAmt"), maxcolnum)

        if maxcolnum:
            yyyy, mm = frompriod//100, frompriod%100
            dy = yyyy - cls.startYear
            dm = mm - cls.startMonth
            startcolumn = dy * 12 + dm + 4

            for col in range(startcolumn, maxcolnum+1):
                for chart in cls.code2chart.values():
                    chart.fillZero(col)
            charts= iter(cls.code2chart.values())
            next(charts).__make_forecast(maxcolnum+1, True)
            for chart in charts:
                chart.__make_forecast(maxcolnum+1)
        if len(datas["items"])==0:
            print("업데이트할 데이터가 없습니다.")

        if to_add:
            print(f"{','.join(cls.code2country.get(i, f'코드없음-{i}') for i in to_add)}의 국가 차트가 존재하지 않습니다. addNation=YES 옵션을 통해 차트를 생성 가능합니다.")

    @classmethod
    def create_chart(cls, code, worksheet):
        '''
        code를 받아서 해당 코드의 국가에 해당하는 표를 만든다.
        양식을 생성해야한다.
        오브젝트를 code2chart[code]=chart와 같이 넣어 코드에 매칭시켜야 한다.
        :param code: 국가코드
        :return: 없음.
        '''

        row = (len(cls.code2chart)-1) * (len(cls.itemInfos) + 4) + 1
        worksheet.cell(row=row, column=1, value=cls.code2country.get(code, code))
        rn = row+3
        for i in range(len(cls.totalIndexes)):
            worksheet.cell(row=rn+i, column=1, value=TOTAL)
        worksheet.cell(row=rn+len(cls.totalIndexes), column=1, value=TOTAL_ALL_BY_COUNTRY)
        chart = cls(worksheet, row)
        cls.code2chart[code]=chart
        for column in range(4, 4+(cls.foreYear+1 - cls.startYear)*12):
            chart.worksheet.cell(row=chart.rownum + 2, column=column, value=column - 3)

    def __fill_chart(self, data, colname):
        '''
        차트를 채우는 함수.
        :param data: request로 불러온 json데이터의 일부.
        data["hsSgn"], data["priodTitle"], data["expTtwg"], data["expUsdAmt"], data["impTtwg"], data["impUsdAmt"], data["cmtrBlncAmt"] 를 통해서 원하는 데이터를 얻을 수 있다.
        :param colname: 채우고자 하는것. 보통은 expUsdAmt 또는 impUsdAMt의 문자열이 들어온다.

        self.code2row 또는 dataChart.code2row : dict. 품목코드 -> 몇번째 줄인가.
        self.rownum: 숫자. 차트가 몇 번째 row부터 시작하는가.
        target cell의 row좌표는 code2row[hsSgn코드] + self.rownum이 된다.

        self.worksheet: target cell이 위치한 엑셀의 워크시트이다.
        '''
        if(data["hsSgn"].isdigit()):
            yyyy, mm = data["priodTitle"].split(".")
            dy = int(yyyy) - self.startYear
            dm = int(mm) - self.startMonth
            targetcolnum = dy * 12 + dm + 4

            # 엑셀 시트에 값이 채워질 때 공백이 포함되는 문제를 해결하기 위한 재검증 코드
            value = 0
            if colname in data:
                try:
                    # 천의 자리 이상의 값은 문자열에 콤마(,)가 포함돼 치환 후 정수로 변환
                    numeric_value = int(data[colname].replace(",", ""))
                    if numeric_value != 0:
                        value = numeric_value
                except (ValueError, TypeError):
                    pass

            row = self.code2row[data["hsSgn"]] + self.rownum
            self.worksheet.cell(row=row, column=targetcolnum, value=value)
            return targetcolnum
        return -1

    def __make_forecast(self, colnum, flag=False):
        '''
        예측하도록 채우는 함수. 데이터가 12월을 채우거나, 오늘의 연도가 넘어갔을 경우 발동한다.
        표의 서식을 따라서 몇년 몇월인지, 몇번째 데이터인지, 그리고 이후 데이터는 모두 forecast 함수를 사용하여 값을 채우도록 만들어진다.
        numToChr 함수를 통해서 엑셀의 열 문자가 뭔지 쉽게 구할 수 있다.
        :param colnum: 해당 열부터 12월까지 채워서 forecast를 만들것. total은 제외.
        :return:
        '''
        endcol = (((colnum - 3)//12) +1) * 12 + 4
        for col in range(colnum, endcol):
            for row in range(len(self.itemInfos)-1):
                if row not in self.totalIndexes:
                    cn=numToChr(col)
                    bn=numToChr(col-1)
                    an = numToChr(col-12)   # 최근 1년 기준
                    nrow=self.rownum+row+3
                    self.worksheet.cell(row=nrow, column=col, value=f"=FORECAST({cn}3, {an}{nrow}:{bn}{nrow}, {an}3:{bn}3)")

        if flag:
            y = self.startYear + (endcol - 4)//12
            for m,col in enumerate(range(endcol-12,endcol)):
                self.worksheet.cell(row=2, column=col, value=f"{y}년 {m+1}월")
                self.worksheet.cell(row=3, column=col, value=col-3)

    @classmethod
    def save(cls):
        if cls.saveNew=="YES":
            try:
                cls.wb.save(f"{cls.fileTempt}.xlsx")
            except:
                num=0
                while num<100:
                    try:
                        cls.wb.save(f"{cls.fileTempt}{num}.xlsx")
                        print(f"{cls.fileTempt}{num}.xlsx 파일에 저장되었습니다.")
                        break
                    except:
                        num+=1
                print("파일 저장에 실패했습니다. 새 파일 경로를 다른 곳에서 참조 중이니 종료하거나 임시파일명을 바꿔주세요.")

        else:
            while True:
                try:
                    cls.wb.save(FILENAME)
                    break
                except:
                    if "y"==input("파일을 어딘가에서 열어서 참조하고 있습니다. 참조를 종료 후 콘솔에 소문자 y를 입력하면 저장 가능합니다."):
                        continue
                    break

def main():
    try:
        load_wb = load_workbook(FILENAME)
    except:
        print(f"{FILENAME} 을 찾을 수 없습니다. 기존 양식 파일을 지정된 위치에 넣어주세요.")
        input("종료하기 위해서 엔터를 눌러주세요")
        sys.exit(0)
    dataChart.settings(load_wb)
    dataChart.create_standard_chart(load_wb[STANDARD_SHEET_NAME], 1)
    dataChart.create_country_chart(load_wb[BYNATION_SHEET_NAME])
    dataChart.run()
    dataChart.save()
    input("작업을 안전하게 종료하였습니다. 종료하기 위해 엔터를 눌러주세요.")

if __name__ == "__main__":
    main()
